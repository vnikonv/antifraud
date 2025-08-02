import streamlit as st
import pandas as pd
import numpy as np
import joblib
import json
import os
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics import classification_report, precision_score, recall_score, roc_auc_score, accuracy_score
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from xgboost import XGBClassifier
from catboost import CatBoostClassifier
from lightgbm import LGBMClassifier
from imblearn.under_sampling import RandomUnderSampler
from sklearn.impute import SimpleImputer
import matplotlib.pyplot as plt
from sklearn.metrics import roc_curve
import time

st.set_page_config(page_title='Antifraud AI', page_icon=':mag:', layout='wide', initial_sidebar_state = 'auto')

st.markdown("""
<style>
/* Общий стиль сайдбара */
section[data-testid="stSidebar"] {
    background-color: #00B2CA !important;
    color: white !important;
    padding-top: 10px !important;
}

/* Надписи и тексты в сайдбаре */
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] div,
section[data-testid="stSidebar"] span {
    color: white !important;
    font-weight: 800 !important;
    font-size: 16px !important;
    border-radius: 12px !important;
}

/* Селектбокс и поля ввода */
section[data-testid="stSidebar"] div[role="combobox"],
section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] select {
    background-color: #000000 !important;
    border: none !important;
    border-radius: 8px !important;
    color: white !important;
    box-shadow: none !important;
}

/* SVG иконки в сайдбаре */
section[data-testid="stSidebar"] svg {
    color: white !important;
}

/* Выпадающие списки и опции */
ul[role="listbox"],
li[role="option"] {
    background-color: #000000 !important;
    color: white !important;
    border-radius: 8px !important;
    border: none !important;
}

li[role="option"]:hover,
li[aria-selected="true"] {
    background-color: #000000 !important;
}

/* Устранение стандартных границ и теней от BaseWeb */
[data-baseweb="select"],
[data-baseweb="select"] * {
    background-color: #000000 !important;
    color: white !important;
    height: auto !important;
    border: none !important;
    box-shadow: none !important;
}

img {
    border-radius: 32px !important;
}



label div p {
    font-size: 24px !important;
    font-weight: 600 !important;
}

[data-testid="stMainBlockContainer"] {
    padding-top: 5% !important;
}

</style>
""", unsafe_allow_html=True)

st.set_page_config(page_title='Antifraud AI', page_icon = "ammit_search.png", layout = 'wide', initial_sidebar_state = 'auto')

icon = "nku_icon.png"

left_co, cent_co, last_co = st.columns([0.35, 0.3, 0.35])
with cent_co:
    st.image("ammit.png")
col1, col2 = st.columns([0.1, 0.9])
with col1:
    st.image(icon, width=100)
with col2:
    st.title("Антифрод ML")

page = st.sidebar.selectbox("Выберите страницу:", ["🏋️‍♂️ Обучение модели", "🔍 Проверка"])

def align_features(df, feature_names):
    for col in feature_names:
        if col not in df.columns:
            df[col] = 0
    return df[feature_names]

# Classifier models selection
models = {
    "RandomForest": RandomForestClassifier(n_estimators=100, class_weight="balanced", random_state=42),
    "GradientBoosting": GradientBoostingClassifier(n_estimators=250, learning_rate=0.1, max_depth=6, random_state=42),
    "CatBoost": CatBoostClassifier(verbose=0, random_state=42),
    "LightGBM": LGBMClassifier(random_state=42),
    "XGBoost": XGBClassifier(use_label_encoder=False, eval_metric="logloss", random_state=42)
}

if page == "🏋️‍♂️ Обучение модели":
    st.header("Обучение модели")
    uploaded_file = st.file_uploader("Загрузите CSV с колонкой 'IsFraud'", type=["csv"])
    use_default = st.checkbox("Или используйте стандартный датасет")

    if use_default:
        uploaded_file = "./account_data.csv"
        st.info("✅ Вы используете встроенный пример датасета")
    

    model_name = st.selectbox("Выберите модель:", list(models.keys()))

    if uploaded_file:
        # Loading data
        df = pd.read_csv(uploaded_file)

        # Выбор переменных с помощью флажков
        selected_features = st.sidebar.multiselect(
            "🔧 Отметьте переменные для отображения:",
            df.columns.tolist(),
            default=df.columns.tolist()
        )
        
        st.dataframe(df[selected_features])

        if st.button("🚀 Обучить модель"):
            # GB Flag = "IsFraud" column, binary classification task
            y = df["IsFraud"].map({"Yes": 1, "No": 0})
            X = df.drop(columns=["IsFraud"])

            # Detect categorical columns and encode them using LabelEncoder as str
            for col in X.select_dtypes(include="object").columns:
                le = LabelEncoder()
                X[col] = le.fit_transform(X[col].astype(str))

            # Balancing samples using RandomUnderSampler
            rus = RandomUnderSampler(random_state=42)
            X_res, y_res = rus.fit_resample(X, y)

            # Train/test split
            X_train, X_test, y_train, y_test = train_test_split(X_res, y_res, test_size=0.2, stratify=y_res, random_state=42)

            # Feature scaling
            scaler = StandardScaler()
            X_train = scaler.fit_transform(X_train)
            X_test = scaler.transform(X_test)

            model = models[model_name]
            start_time = time.time()
            model.fit(X_train, y_train)
            train_time = time.time() - start_time

            y_pred = model.predict(X_test)
            y_proba = model.predict_proba(X_test)[:, 1] if hasattr(model, "predict_proba") else None

            accuracy = accuracy_score(y_test, y_pred)
            precision = precision_score(y_test, y_pred)
            recall = recall_score(y_test, y_pred)
            roc_auc = roc_auc_score(y_test, y_proba) if y_proba is not None else 0.0
            gini = 2 * roc_auc - 1
            def ks_statistic(y_true, y_proba):
                from scipy.stats import ks_2samp
                return ks_2samp(y_proba[y_true == 1], y_proba[y_true == 0]).statistic
            ks = ks_statistic(y_test.to_numpy(), y_proba) if y_proba is not None else 0.0

            st.success(f"✅ {model_name} модель обучена и сохранена")

            st.subheader("📈 Результаты обучения (тест)")
            st.markdown(f"⏱ **Время обучения:** {train_time:.2f} сек")
            st.markdown(f"📊 **Gini индекс:** {gini:.4f}")
            st.markdown(f"📊 **KS индекс:** {ks:.4f}")
            st.markdown(f"✅ **Accuracy:** {accuracy:.4f}")
            st.markdown(f"🎯 **Precision:** {precision:.4f}")
            st.markdown(f"🔁 **Recall:** {recall:.4f}")

            y_train_pred = model.predict(X_train)
            y_train_proba = model.predict_proba(X_train)[:, 1] if hasattr(model, "predict_proba") else None
            accuracy_train = accuracy_score(y_train, y_train_pred)
            precision_train = precision_score(y_train, y_train_pred)
            recall_train = recall_score(y_train, y_train_pred)
            roc_auc_train = roc_auc_score(y_train, y_train_proba) if y_train_proba is not None else 0.0
            gini_train = 2 * roc_auc_train - 1
            ks_train = ks_statistic(y_train.to_numpy(), y_train_proba) if y_train_proba is not None else 0.0

            st.subheader("📈 Результаты обучения (обучение)")
            st.markdown(f"⏱ **Время обучения:** {train_time:.2f} сек")
            st.markdown(f"📊 **Gini индекс:** {gini_train:.4f}")
            st.markdown(f"📊 **KS индекс:** {ks_train:.4f}")
            st.markdown(f"✅ **Accuracy:** {accuracy_train:.4f}")
            st.markdown(f"🎯 **Precision:** {precision_train:.4f}")
            st.markdown(f"🔁 **Recall:** {recall_train:.4f}")

            st.subheader("🔢 Пример расчётов")
            metrics_df = pd.DataFrame({
                "Метрика": ["Время обучения (сек)", "Gini", "KS", "Accuracy", "Precision", "Recall", "ROC AUC"],
                "Значение": [round(train_time, 2), round(gini, 4), round(ks, 4), round(accuracy, 4),
                             round(precision, 4), round(recall, 4), round(roc_auc, 4)]
            })
            metrics_df_train = pd.DataFrame({
                "Метрика": ["Время обучения (сек)", "Gini", "KS", "Accuracy", "Precision", "Recall", "ROC AUC"],
                "Значение": [round(train_time, 2), round(gini_train, 4), round(ks_train, 4), round(accuracy_train, 4),
                             round(precision_train, 4), round(recall_train, 4), round(roc_auc_train, 4)]
            })
            st.subheader("📊 Сводная таблица метрик (тест)")
            st.table(metrics_df)

            st.subheader("📊 Сводная таблица метрик (обучение)")
            st.table(metrics_df_train)

            fig_roc = plt.figure()
            fpr, tpr, _ = roc_curve(y_test, y_proba)
            plt.plot(fpr, tpr, label=f"ROC AUC = {roc_auc:.2f}")
            plt.plot([0, 1], [0, 1], linestyle="--", color="gray")
            plt.xlabel("False Positive Rate")
            plt.ylabel("True Positive Rate")
            plt.title("ROC-кривая")
            plt.legend()
            st.pyplot(fig_roc)

            st.subheader("🔍 Распределение вероятностей")
            fig_hist = plt.figure()
            plt.hist(y_proba[y_test == 0], bins=30, alpha=0.6, label="Класс 0 (не мошенник)")
            plt.hist(y_proba[y_test == 1], bins=30, alpha=0.6, label="Класс 1 (мошенник)")
            plt.title("📊 Распределение вероятностей по классам")
            plt.xlabel("Предсказанная вероятность")
            plt.ylabel("Количество")
            plt.legend()
            st.pyplot(fig_hist)
            st.markdown(f"**Рассчитанный балл (пример):** {y_pred[0]}")
            st.markdown(f"**Вероятностная оценка (пример):** {y_proba[0]:.4f}")

            st.subheader("📌 Использованные переменные")
            st.code(", ".join(X.columns))

            subset_info = pd.DataFrame({
                "target": pd.concat([y_train, y_test], ignore_index=True),
                "subset": ["train"] * len(y_train) + ["test"] * len(y_test)
            })
            st.subheader("🧪 Разметка выборок")
            st.dataframe(subset_info.sample(10, random_state=42))

            X_test_raw = pd.DataFrame(scaler.inverse_transform(X_test), columns=X.columns)
            results_df = X_test_raw.copy()
            results_df["Предсказанный класс"] = y_pred
            results_df["Вероятность"] = y_proba
            results_df["Истинный IsFraud"] = y_test.values

            csv = results_df.to_csv(index=False).encode("utf-8")
            col_csv, col_xls = st.columns(2)
            with col_csv:
                st.download_button(
                    label="📥 Скачать результаты на тестовой выборке в CSV",
                    data=csv,
                    file_name="test_predictions.csv",
                    mime="text/csv"
                )
            with col_xls:
                import io
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Test Predictions"
                for col_num, column in enumerate(results_df.columns, 1):
                    ws.cell(row=1, column=col_num, value=column)
                for row_num, row in enumerate(results_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                st.download_button(
                    label="📥 Скачать результаты на тестовой выборке в XLS",
                    data=excel_buffer.getvalue(),
                    file_name="test_predictions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            st.success(f"✅ {model_name} модель сохранена")

            os.makedirs("models", exist_ok=True)
            
            # Save model
            joblib.dump(model, f"models/{model_name}_model.pkl")

                # Save scaler
            joblib.dump(scaler, f"models/{model_name}_scaler.pkl")

            # Save feature names
            with open(f"models/{model_name}_features.json", "w") as f:
                json.dump(X.columns.tolist(), f)

if page == "🔍 Проверка":
    # Scan all models that have a saved _model.pkl file
    available_models = [
    f.replace("_model.pkl", "")
    for f in os.listdir("models")
    if f.endswith("_model.pkl")
    ]

    if available_models:
        selected_model = st.selectbox("Выберите модель для предсказания:", available_models)

        model = joblib.load(f"models/{selected_model}_model.pkl")
        scaler = joblib.load(f"models/{selected_model}_scaler.pkl")

        with open(f"models/{selected_model}_features.json", "r") as f:

            feature_names = json.load(f)

        uploaded_file = st.file_uploader("Загрузите тестовый файл", type=["csv", "xlsx", "xls"])

        if uploaded_file:
            df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
            df_original = df.copy()

            # Align features (implement align_features if needed)
            df_aligned = align_features(df, feature_names)
            # Encode categorical columns
            for col in df_aligned.select_dtypes(include="object").columns:
                le = LabelEncoder()
                df_aligned[col] = le.fit_transform(df_aligned[col].astype(str))
            imputer = SimpleImputer(strategy="mean")
            df_imputed = pd.DataFrame(imputer.fit_transform(df_aligned), columns=df_aligned.columns)
            df_scaled = pd.DataFrame(scaler.transform(df_imputed), columns=df_imputed.columns)

            predictions = model.predict_proba(df_scaled)[:, 1]
            is_fraud_pred = model.predict(df_scaled)
            # Create output DataFrame with only index, Вероятность мошенничества, and IsFraud
            output_df = pd.DataFrame({
                "Вероятность мошенничества": predictions,
                "IsFraud": is_fraud_pred
            })
            st.subheader("📋 Предсказания")
            rows_to_display = st.slider(
                "Количество отображаемых строк",
                min_value=0,
                max_value=min(100, len(output_df)),
                value=min(10, len(output_df)),
                step=1
            )
            st.dataframe(output_df.head(rows_to_display))
            
            # Two download buttons: CSV and XLS
            col_csv, col_xls = st.columns(2)
            with col_csv:
                csv_pred = output_df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="📥 Скачать предсказания в CSV",
                    data=csv_pred,
                    file_name="predictions.csv",
                    mime="text/csv"
                )
            with col_xls:
                import io
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Predictions"
                for col_num, column in enumerate(output_df.columns, 1):
                    ws.cell(row=1, column=col_num, value=column)
                for row_num, row in enumerate(output_df.values, 2):
                    for col_num, value in enumerate(row, 1):
                        ws.cell(row=row_num, column=col_num, value=value)
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                st.download_button(
                    label="📥 Скачать предсказания в XLS",
                    data=excel_buffer.getvalue(),
                    file_name="predictions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
           
    else:
        st.error("❌ Сначала обучите модель")